import shutil # for copying files
import openpyxl # for reading and writing .xlsx excel files
import pandas as pd # for reading and writing .xls excel files
from datetime import datetime
import configparser
import warnings
import os

def saveOneFactory():
    #### local configs
    configParser = configparser.RawConfigParser()
    configParser.read('folderConfig.txt')
    archiveFolder = configParser.get('folderPath', 'savePath')
    oneFactorySavePath = configParser.get('folderPath', 'savePathOneFactory')
    configPath = configParser.get('folderPath', 'configPath')
    title = configParser.get('folderPath', 'title') # one factory excel doc title
    CMM = configParser.get('folderPath', 'CMM') # Test machine identifier
    folder = 'temp/' # source data

    for count, filename in enumerate(os.listdir(folder)):

        newFilename = filename.split('.')  # extract date and time from file name
        newFilename = newFilename[0] # the new filename
        path =oneFactorySavePath + newFilename + '.xls'
        shutil.copy('OneFactory_Template.xls', path) # copy template and rename
        df = pd.read_excel(path) # read copied template

        ##now = datetime.now() # current date and time
        ##df.loc[3,'Unnamed: 3'] = now.strftime('%m/%d/%Y') # Date
        ##df.loc[6,'Unnamed: 3'] = now.strftime('%H:%M') # Time
        
        partNumber_dateTime = newFilename.split('(')
        partNumber = partNumber_dateTime[0]
        dateTime = partNumber_dateTime[1]
        dateTime = dateTime.replace(')', '')
        dateTime = dateTime.split('-')
        date = dateTime[0]
        time = dateTime[1]
        time = [time[i:i+4] for i in range(0, len(date), 4)]
        time = time[0]
        time = [time[i:i+2] for i in range(0, len(date), 2)]
        time = time[0] + ':' + time[1]
        year_MonthDay = [date[i:i+4] for i in range(0, len(date), 4)]
        monthDay = year_MonthDay[1]
        month_Day = [monthDay[i:i+2] for i in range(0, len(monthDay), 2)]
        month = month_Day[0]
        day = month_Day[1]
        year = year_MonthDay[0]
        date = month + '/' + day + '/' + year
        df.loc[3,'Unnamed: 3'] = date # Date
        df.loc[6,'Unnamed: 3'] = time # Time
        #print(newFilename, date, time)
        
        if 'Failure' in partNumber:
            partNumber = partNumber.replace('Failure-','')
        partNumber = partNumber.split('-')
        partNumber = partNumber[0] + '-' + partNumber[1] + '-' + partNumber[2]
        df.loc[6,'Unnamed: 5'] = partNumber # Part Number

        #### local configs
        df.loc[0,'Unnamed: 1'] = title # Header
        df.loc[9,'Unnamed: 1'] = os.getlogin() # operator
        df.loc[9,'Unnamed: 3'] = CMM # CMM
        
        #### Motor Configs
        configFilePath = configPath+'/'+partNumber+'.txt'
        configParser.read(configFilePath)
        testPlan = configParser.get(partNumber, 'testPlan')
        df.loc[3,'Unnamed: 1'] = testPlan # Measurment Plan
        poleStrengthMinimum = configParser.get(partNumber, 'poleStrengthMinimum')
        poleStrengthMaximum = configParser.get(partNumber, 'poleStrengthMaximum')
        mcrFreq = configParser.get(partNumber, 'mcrFreq')
        resMin = configParser.get(partNumber, 'resMin')
        resMax = configParser.get(partNumber, 'resMax')
        indMin = configParser.get(partNumber, 'indMin')
        indMax = configParser.get(partNumber, 'indMax')
        mode = configParser.get(partNumber, 'mode')
        gptV = configParser.get(partNumber, 'gptV')
        gptI = configParser.get(partNumber, 'gptI')
        gptImin = configParser.get(partNumber, 'gptImin')
        gptFreq = configParser.get(partNumber, 'gptFreq')

        dataframe = openpyxl.load_workbook(folder + filename)# Load the dataframe (Call out absolute path)

        try:
            sheetOne = dataframe['Flux'] # Read SheetOne
            i=14 # starting row
            j=1 # pole
            #### Polarity Tests
            for row in range(1, sheetOne.max_row): # Iterate the loop to read the cell values startin row to ending row
                for col in sheetOne.iter_cols(3,3): # iterate through all columns = (sheetOne.max_column): (starting col, to end col)
                    nominalVal = col[row].value
                    if nominalVal == 'N':
                        nominalVal = 1
                    else:
                        nominalVal = -1
                    df.loc[i,'Unnamed: 0'] = '#' + str(i - 13) + '_Polarity_Test_' + str(j) # Characteristics
                    df.loc[i,'Unnamed: 2'] = int(nominalVal) # Nominal values
                    df.loc[i,'Unnamed: 3'] = 0 # Upper Tol values
                    df.loc[i,'Unnamed: 4'] = 0 # Lower Tol values
                for col in sheetOne.iter_cols(6,6): # iterate through all columns = (sheetOne.max_column): (starting col, to end col)
                    actualVal = col[row].value
                    if actualVal == 'N':
                        actualVal = 1
                    else:
                        actualVal = -1
                    df.loc[i,'Unnamed: 1'] = actualVal # Actual values
                    df.loc[i,'Unnamed: 5'] = actualVal - nominalVal # Deviation values
                    i+=1 
                    j+=1

            #### Strength Tests
            j=1  # pole
            poleStrengthMinimum = float(poleStrengthMinimum)
            poleStrengthMaximum = float(poleStrengthMaximum)
            nominalVal = poleStrengthMaximum - poleStrengthMinimum
            poleStrengthMaximum = poleStrengthMaximum - nominalVal
            poleStrengthMinimum = poleStrengthMinimum - nominalVal
            for row in range(1, sheetOne.max_row): # Iterate the loop to read the cell values startin row to ending row
                for col in sheetOne.iter_cols(4,4): # iterate through all columns = (sheetOne.max_column): (starting col, to end col)
                    actualVal = float(col[row].value)
                    df.loc[i,'Unnamed: 0'] = '#' + str(i - 13) + '_Flux_Strenth_(mT)_' + str(j) # Characteristics
                    df.loc[i,'Unnamed: 1'] = actualVal # Actual values
                    df.loc[i,'Unnamed: 2'] = nominalVal # Nominal values
                    df.loc[i,'Unnamed: 3'] = poleStrengthMaximum # Upper Tol values
                    df.loc[i,'Unnamed: 4'] = poleStrengthMinimum # Lower Tol values
                    df.loc[i,'Unnamed: 5'] = actualVal - nominalVal # Deviation values
                    i+=1 
                    j+=1
        except:
            pass
        #### Hi-Pot Testing
        try:
            sheetTwo = dataframe['HiPot'] # Read SheetTwo
            gptI = float(gptI)
            gptImin = float(gptImin)
            nominalVal = gptI - gptImin
            nominalVal = nominalVal / 2
            gptI = gptI - nominalVal
            gptImin = gptImin - nominalVal
            for row in range(1, sheetTwo.max_row): # Iterate the loop to read the cell values starting row to ending row
                for col in sheetTwo.iter_cols(3,3): # iterate through all columns = (sheetOne.max_column): (starting col, to end col)
                    voltage = col[row].value
                for col in sheetTwo.iter_cols(4,4): # iterate through all columns = (sheetOne.max_column): (starting col, to end col)
                    actualVal = str(col[row].value)
                    actualVal = actualVal.split(' ')
                    actualVal = actualVal[0]
                    df.loc[i,'Unnamed: 0'] = '#' + str(i - 13) + '_HiPot(mA)_' + mode +'_'+ voltage +'_'+ gptFreq+'hz'# Characteristics
                    df.loc[i,'Unnamed: 1'] = float(actualVal) # Actual values
                    df.loc[i,'Unnamed: 2'] = nominalVal # Nominal values
                    df.loc[i,'Unnamed: 3'] = gptI # Upper Tol values
                    df.loc[i,'Unnamed: 4'] = gptImin # Lower Tol values
                    df.loc[i,'Unnamed: 5'] = float(actualVal) - nominalVal # Deviation values
                    i+=1
        except:
            pass
        try:
            sheetThree = dataframe['IndRes'] # Read SheetThree
            testID = ['_A-B_', '_A-C_', '_B-C_']
            #### Inductance Testing
            j=0
            indMax = float(indMax)
            indMin = float(indMin)
            nominalVal = indMax + indMin
            nominalVal = nominalVal / 2
            indMax = indMax - nominalVal
            indMin = indMin - nominalVal
            for row in range(1, sheetThree.max_row): # Iterate the loop to read the cell values
                for col in sheetThree.iter_cols(3,3): # iterate through all columns = (sheetOne.max_column): (starting col, to end col)
                    actualVal = float(col[row].value)
                    df.loc[i,'Unnamed: 0'] = '#' + str(i - 13) + testID[j] + 'Inductance(mH)' # Characteristics
                    df.loc[i,'Unnamed: 1'] = actualVal # Actual values
                    df.loc[i,'Unnamed: 2'] = nominalVal # Nominal values
                    df.loc[i,'Unnamed: 3'] = indMax # Upper Tol values
                    df.loc[i,'Unnamed: 4'] = indMin # Lower Tol values
                    df.loc[i,'Unnamed: 5'] = actualVal - nominalVal # Deviation values
                    i+=1
                    j+=1

            #### Resistance Testing
            j=0
            resMax = float(resMax)
            resMin = float(resMin)
            nominalVal = resMax + resMin
            nominalVal = nominalVal / 2
            resMax = resMax - nominalVal
            resMin = resMin - nominalVal
            for row in range(1, sheetThree.max_row): # Iterate the loop to read the cell values
                for col in sheetThree.iter_cols(5,5): # iterate through all columns = (sheetOne.max_column): (starting col, to end col)
                    actualVal = float(col[row].value)
                    df.loc[i,'Unnamed: 0'] = '#' + str(i - 13) + testID[j] + 'Resistance(mΩ)' # Characteristics
                    df.loc[i,'Unnamed: 1'] = actualVal # Actual values
                    df.loc[i,'Unnamed: 2'] = nominalVal # Nominal values
                    df.loc[i,'Unnamed: 3'] = resMax # Upper Tol values
                    df.loc[i,'Unnamed: 4'] = resMin # Lower Tol values
                    df.loc[i,'Unnamed: 5'] = actualVal - nominalVal # Deviation values
                    i+=1
                    j+=1
        except:
            pass
        warnings.filterwarnings("ignore")
        df.to_excel(path, index=False, header=False, sheet_name='Report')
        shutil.move(folder + filename, archiveFolder)

        print(filename + " appended successfully.") 
        #print(df.to_string())

    print("All data appended successfully.") 
